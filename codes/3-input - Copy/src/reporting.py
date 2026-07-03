from __future__ import annotations
import csv
import json
import os
from typing import Dict, Any, List, Optional


def ensure_dir(path: str) -> None:
    if path:
        os.makedirs(path, exist_ok=True)


def write_csv(path: str, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    ensure_dir(os.path.dirname(path))

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_json(path: str, data: Dict[str, Any]) -> None:
    ensure_dir(os.path.dirname(path))

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def write_markdown(path: str, content: str) -> None:
    ensure_dir(os.path.dirname(path))

    with open(path, "w", encoding="utf-8") as f:
        f.write(content)


def markdown_table(headers: List[str], rows: List[List[Any]]) -> str:
    out = []

    out.append("| " + " | ".join(headers) + " |")
    out.append("| " + " | ".join(["---" for _ in headers]) + " |")

    for row in rows:
        out.append("| " + " | ".join(str(x) for x in row) + " |")

    return "\n".join(out)


def safe_ratio(a: float, b: float) -> float:
    if b == 0:
        return float("inf")
    return float(a) / float(b)
def fmt_ratio(a: float, b: float) -> str:
    r = safe_ratio(a, b)

    if r == float("inf"):
        return "inf"

    return f"{r:.2f}×"


def percent(x: float) -> str:
    return f"{100.0 * x:.2f}%"


def write_xlsx_workbook(
    path: str,
    sheets: Dict[str, List[Dict[str, Any]]],
) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    ensure_dir(os.path.dirname(path))

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    header_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True)
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        if not rows:
            ws["A1"] = "No data"
            continue

        headers = list(rows[0].keys())

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
                cell.border = border

        for col_idx, h in enumerate(headers, start=1):
            width = max(len(str(h)) + 2, 12)
            for row_idx in range(2, len(rows) + 2):
                value = ws.cell(row=row_idx, column=col_idx).value
                width = max(width, min(len(str(value)) + 2, 35))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

    wb.save(path)
    return path